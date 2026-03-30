import openpyxl

CF_INPUT_FILE = r"H:\My Drive\3.PDC\Crop planning CycleFarm.xlsx"

# Keys that signal the start of a new crop entry (row[0] is the name, row[1] is None)
PROPERTY_KEYS = {"Pépinière", "itinéraire", "Séries", "Variété", "Fournisseur",
                 "Quantité ", "Récolte auto-cueillette", "Commentaire"}


def is_crop_header(row: tuple) -> bool:
    """A crop header row has a name in col A and nothing in col B."""
    return (
        row[0] is not None
        and row[1] is None
        and row[0] not in PROPERTY_KEYS
    )


def parse_sheet(ws) -> list[dict]:
    """
    Parse one sheet and return a list of crop dicts.
    The sheet name is used as the 'category' value.
    """
    category = ws.title
    rows = [row for row in ws.iter_rows(values_only=True)]

    crops: list[dict] = []
    current_crop: dict | None = None
    current_key: str | None = None

    for idx, row in enumerate(rows):
        if idx < 1:
            continue

        col_a, col_b = row[0], row[1]

        # New crop header
        if is_crop_header(row):
            if current_crop is not None:
                crops.append(current_crop)
            current_crop = {"name": col_a, "category": category}
            current_key = None
            continue

        if current_crop is None:
            continue

        # Continuation line (col A is None, col B has a value)
        if col_a is None and col_b is not None and current_key:
            existing = current_crop.get(current_key)
            if isinstance(existing, list):
                existing.append(col_b)
            else:
                current_crop[current_key] = [existing, col_b]
            continue

        # Normal property row
        if col_a in PROPERTY_KEYS and col_b is not None:
            current_key = col_a.strip()
            current_crop[current_key] = col_b
        elif col_a in PROPERTY_KEYS and col_b is None:
            current_key = col_a.strip()
            # value may come in a continuation row
        elif col_a is None and col_b is None:
            # blank row – keep current_key so multi-line values still work
            pass

    # Don't forget the last crop
    if current_crop is not None:
        crops.append(current_crop)

    return [item for item in crops if "Séries" in item]


def parse_workbook(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    all_crops: list[dict] = []
    for idx, sheet_name in enumerate(wb.sheetnames):
        if idx > 7:
            all_crops.extend(parse_sheet(wb[sheet_name]))
    return all_crops


def parse_itk():
    return parse_workbook(CF_INPUT_FILE)


if __name__ == "__main__":
    parse_itk()
