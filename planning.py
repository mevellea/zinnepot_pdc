import json
from pathlib import Path
from collections import defaultdict
from typing import List

from jinja2 import Template
from datetime import datetime

from openpyxl import load_workbook
from crop import Crop, load_crops
import copy

from customers import gen_customer_html

INPUT_FILE = r"H:\My Drive\3.PDC\PDC.xlsx"
ITK_OUTPUT_FILE = r"H:\My Drive\3.PDC\export\itk.html"
TASKS_OUTPUT_FILE = r"H:\My Drive\3.PDC\export\taches.html"
CAL_OUTPUT_FILE = r"H:\My Drive\3.PDC\export\calendar.html"


class CropImplantation(Crop):
    block: int
    garden: int
    bed: int

    variety: str

    grow_start: int
    grow_end: int
    harvest_start: int
    harvest_end: int

    sowing_week: int
    sowing_done: bool
    transplanting_done: bool

    def __init__(self, **entries):
        super().__init__(**entries)
        self.variety = ""
        if self.harvest_start is None or self.harvest_end is None:
            raise ValueError(f"{self.get_location()} {self.crop}: harvest_start and harvest_end must be set")

    def get_location(self):
        return f"{self.block}-{self.garden}-{self.bed}"

    def update(self, **entries):
        for k, v in entries.items():
            setattr(self, k, v)
        jours_en_pep = getattr(self, "Jours en pep")
        if jours_en_pep == "":
            jours_en_pep = 0
        self.sowing_week = self.grow_start - int(float(jours_en_pep) / 7)

    def __str__(self):
        val = f"{self.get_location()} {self.crop} "
        if self.variety:
            val += f"/ {self.variety} "
        val += f"S={self.sowing_week} DTM=[{self.grow_start}-{self.grow_end}] R=[{self.harvest_start}-{self.harvest_end}]"
        return val


def extract_planning(crops_database) -> List[CropImplantation]:
    print("Parsing planning...")
    wb = load_workbook(INPUT_FILE, data_only=True)
    wb.active = wb['PDC']
    ws = wb.active

    weeks = {col: ws.cell(row=1, column=col).value for col in range(4, ws.max_column + 1)}

    def has_thick_border(cell):
        b = cell.border
        return any(
            s and s.style in ("medium", "thick")
            for s in [b.left, b.right, b.top, b.bottom]
        )

    def cell_color(cell):
        fill = cell.fill
        if fill and fill.start_color:
            return fill.start_color.rgb
        return None

    bloc = jardin = planche = None
    crops_implantations = []

    for row in range(3, ws.max_row + 1):

        if ws.cell(row=row, column=1).value:
            bloc = ws.cell(row=row, column=1).value
        if ws.cell(row=row, column=2).value:
            jardin = ws.cell(row=row, column=2).value
        if ws.cell(row=row, column=3).value:
            planche = ws.cell(row=row, column=3).value

        col = 4

        while col <= ws.max_column:

            cell = ws.cell(row=row, column=col)

            if not has_thick_border(cell):
                col += 1
                continue

            start_col = col
            segment_color = cell_color(cell)

            crop = None

            while col <= ws.max_column:
                c = ws.cell(row=row, column=col)

                if cell_color(c) != segment_color:
                    break

                if crop is not None and c.value == "R":
                    break

                if c.value and c.value != "R":
                    crop = c.value

                col += 1

            grow_start = weeks[start_col]
            grow_end = weeks[col - 1]

            # ----- SEGMENT RECOLTE -----
            harvest_start = None
            harvest_end = None

            while col <= ws.max_column:
                c = ws.cell(row=row, column=col)

                if c.value != "R":
                    break

                if harvest_start is None:
                    harvest_start = weeks[col]

                harvest_end = weeks[col]
                col += 1

            if crop:
                sowing_done = "#" in crop
                transplanting_done = "!" in crop
                crop = crop.replace("#", "").replace("!", "").lstrip()

                if " - " in crop:
                    crop_name, variety = crop.split(" - ")
                else:
                    crop_name, variety = crop, ""

                try:
                    ref_crop = next(c for c in crops_database if c.crop == crop_name)
                except StopIteration:
                    raise ValueError(f"{crop}: {crop_name} not found")

                crop_impl = copy.deepcopy(ref_crop)
                crop_impl.__class__ = CropImplantation
                crop_impl.update(
                    block=int(bloc),
                    garden=int(jardin),
                    bed=int(planche),
                    grow_start=grow_start,
                    grow_end=grow_end,
                    harvest_start=harvest_start,
                    harvest_end=harvest_end,
                    variety=variety,
                    sowing_done=sowing_done,
                    transplanting_done=transplanting_done
                )
                crops_implantations.append(crop_impl)

            col += 1
    crops_implantations_sorted = reorder_by_int_attr(crops_implantations, "sowing_week")
    return crops_implantations_sorted


def generate_cal_html(harvest: List[CropImplantation], output_file="calendar.html"):
    current_week = datetime.now().isocalendar()[1]
    all_weeks = set()

    for c in harvest:
        all_weeks.update(range(c.grow_start, c.grow_end + 1))
        all_weeks.update(range(c.harvest_start, c.harvest_end + 1))

    weeks = sorted(all_weeks)
    matrix = []

    for c in harvest:
        row = []
        for w in weeks:
            if c.harvest_start <= w <= c.harvest_end:
                row.append("R")
            elif c.grow_start <= w <= c.grow_end:
                row.append("G")
            else:
                row.append("")

        matrix.append({
            "label": f"{c.block}-{c.garden}-{c.bed}",
            "block": c.block,
            "garden": c.garden,
            "bed": c.bed,
            "weeks": row
        })

    # render HTML
    template_file ="templates/template_cal.html"
    with open(template_file, encoding="utf-8") as f:
        template = Template(f.read())

    data = [
        {
            "block": c.block,
            "garden": c.garden,
            "bed": c.bed,
            "label": c.crop,
            "grow_start": c.grow_start,
            "grow_end": c.grow_end,
            "harvest_start": c.harvest_start,
            "harvest_end": c.harvest_end
        }
        for c in harvest
    ]
    data = sorted(
        data,
        key=lambda x: (x["block"], x["garden"], x["bed"])
    )
    html = template.render(
        data=data,
        weeks=weeks,
        current_week=current_week
    )

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"HTML généré : {output_file}")

def extract_harvest(crops_implantations):
    harvest = defaultdict(list)
    for week in range(1, 52):
        harvest[week] = []
    for crop_implantation in crops_implantations:
        for week in range(crop_implantation.harvest_start, crop_implantation.harvest_end):
            if crop_implantation.crop not in harvest[week]:
                harvest[week].append(crop_implantation.crop)
    return {k: v for k, v in harvest.items() if v}


def reorder_by_int_attr(objects, attr, reverse=False):
    return sorted(objects, key=lambda obj: getattr(obj, attr), reverse=reverse)


def generate_html(html_data, template, filename, title):
    template = Path("templates/" + template).read_text(encoding="utf8")
    data = [c.to_dict() for c in html_data]
    json_data = json.dumps(data, ensure_ascii=False)
    html = template.replace("__DATA__", json_data).replace("__TITLE__", title)
    Path(filename).write_text(html, encoding="utf8")


def main():
    crops_database = load_crops()
    generate_html(html_data=crops_database, template="template_itk.html", filename=ITK_OUTPUT_FILE, title="Itinéraires techniques")
    gen_customer_html()

    crops_implantations = extract_planning(crops_database)
    generate_html(html_data=crops_implantations, template="template_tasks.html", filename=TASKS_OUTPUT_FILE, title="Taches")
    generate_cal_html(crops_implantations, output_file=CAL_OUTPUT_FILE)

    harvest = extract_harvest(crops_implantations)
    for week in harvest:
        print(f" w{week} : {", ".join(harvest[week])}")


if __name__ == "__main__":
    main()

